from __future__ import annotations

import os
import platform
import subprocess
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

pytestmark = pytest.mark.integration


def _wait_for_file(path: Path, timeout_seconds: float = 20.0) -> bool:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        if path.exists():
            return True
        time.sleep(0.25)
    return False


def _launchservices_restricted() -> bool:
    probe = subprocess.run(
        ["open", "/System/Applications/TextEdit.app"],
        check=False,
        capture_output=True,
        text=True,
    )
    return "kLSNoExecutableErr" in probe.stderr


@pytest.mark.skipif(platform.system() != "Darwin", reason="macOS-only integration test")
def test_open_with_app_converts_csv(tmp_path: Path):
    if os.environ.get("CSVSAFE_RUN_GUI_TESTS") != "1":
        pytest.skip("Set CSVSAFE_RUN_GUI_TESTS=1 to run GUI integration tests")
    if _launchservices_restricted():
        pytest.skip("LaunchServices app launching is restricted in this environment")

    app_path = Path("dist/CSVSafe.app").resolve()
    if not app_path.exists():
        pytest.fail("dist/CSVSafe.app not found. Run `make build` first.")

    csv_path = tmp_path / "finder_input.csv"
    csv_path.write_text("name,id\nalice,001\n", encoding="utf-8")
    xlsx_path = tmp_path / "finder_input.xlsx"

    proc = subprocess.run(
        ["open", "-n", "-a", str(app_path), str(csv_path)],
        check=False,
        capture_output=True,
        text=True,
    )

    if proc.returncode != 0:
        raise AssertionError(f"open command failed: {proc.stderr.strip()}")

    if not _wait_for_file(xlsx_path):
        raise AssertionError("Expected XLSX output was not created after opening file with app")

    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws["A2"].value == "alice"
    assert ws["B2"].value == "001"
    assert ws["B2"].number_format == "@"
