from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from csvsafe.core.writer import InputSource, write_and_open


def _wb() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "id"])
    ws.append(["alice", "001"])
    return wb


def test_file_source_writes_alongside_and_increments(tmp_path: Path, monkeypatch):
    opened: list[Path] = []
    monkeypatch.setattr("csvsafe.platform.macos.open_file", lambda p: opened.append(Path(p)))

    source = tmp_path / "input.csv"
    source.write_text("a,b\n1,2\n", encoding="utf-8")

    out1 = write_and_open(_wb(), InputSource(origin="file", source_path=source))
    out2 = write_and_open(_wb(), InputSource(origin="file", source_path=source))

    assert out1.name == "input.xlsx"
    assert out2.name == "input (1).xlsx"
    assert out1.exists() and out2.exists()
    assert opened == [out1, out2]


def test_clipboard_output_readonly_and_valid(tmp_path: Path, monkeypatch):
    opened: list[Path] = []
    monkeypatch.setattr("csvsafe.platform.macos.open_file", lambda p: opened.append(Path(p)))

    explicit = tmp_path / "clip.xlsx"
    out = write_and_open(_wb(), InputSource(origin="clipboard"), output_path=explicit)

    assert out == explicit
    assert opened == [explicit]
    loaded = load_workbook(explicit)
    assert loaded.active["A2"].value == "alice"
