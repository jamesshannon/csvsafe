"""Convert parsed rows into XLSX workbook."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from csvsafe.core.parser import ParseResult

MAX_COLUMN_WIDTH = 50


def _sanitize_sheet_name(sheet_name: str) -> str:
    invalid_chars = set('[]:*?/\\')
    sanitized = "".join("_" if ch in invalid_chars else ch for ch in sheet_name.strip())
    if not sanitized:
        sanitized = "Sheet1"
    return sanitized[:31]


def convert_to_workbook(parse_result: ParseResult, sheet_name: str = "Sheet1") -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _sanitize_sheet_name(sheet_name)

    column_lengths: dict[int, int] = {}
    row_count = 0

    for row_count, row in enumerate(parse_result.rows, start=1):
        cleaned = ["" if value is None else str(value) for value in row]
        worksheet.append(cleaned)

        for col_idx, value in enumerate(cleaned, start=1):
            current_len = len(value)
            previous_max = column_lengths.get(col_idx, 0)
            if current_len > previous_max:
                column_lengths[col_idx] = current_len

        for cell in worksheet[row_count]:
            cell.number_format = "@"

        if parse_result.has_header and row_count == 1:
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

    if parse_result.has_header and row_count >= 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, max_len in column_lengths.items():
        width = min(max(max_len + 2, 2), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    return workbook
